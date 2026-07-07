#!/usr/bin/env python3
"""生意参谋 Excel 审计与提取工具模板。

本文件提供可复用且可测试的基础能力：区间解析、日期转换、表头读取和
工作簿审计。业务 Sheet 到 Dashboard JSON 的映射必须根据实际表头实现，
不要在未检查输入时假设固定 Sheet 名或固定列顺序。

用法：
  python extract_template.py --self-test
  python extract_template.py input.xlsx --audit-json audit.json
"""

from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.utils.datetime import from_excel


MISSING_TOKENS = {"", "-", "--", "n/a", "na", "null", "none"}
RANGE_SEPARATORS = r"(?:~|～|至|—|–)"
NUMBER = r"(?:\d+(?:\.\d+)?|\.\d+)"
UNIT = r"(?:万|亿)?"


class ParseError(ValueError):
    """表示非空指标值无法按已知格式解析。"""


def _number_with_unit(number: str, unit: str) -> float:
    multiplier = {"": 1, "万": 10_000, "亿": 100_000_000}[unit]
    return float(number) * multiplier


def _result(low: float, high: float, *, exact: bool = False) -> dict[str, Any]:
    if not math.isfinite(low) or not math.isfinite(high) or low > high:
        raise ParseError(f"非法区间: {low} ~ {high}")
    result: dict[str, Any] = {
        "low": low,
        "mid": (low + high) / 2,
        "high": high,
    }
    if exact:
        result["exact"] = True
    return result


def parse_range(value: Any) -> dict[str, Any] | None:
    """解析生意参谋数值；缺失返回 None，未知非空格式抛出 ParseError。"""
    if value is None:
        return None
    if isinstance(value, bool):
        raise ParseError(f"布尔值不是有效指标: {value!r}")
    if isinstance(value, (int, float)):
        number = float(value)
        if not math.isfinite(number):
            raise ParseError(f"非有限数值: {value!r}")
        return _result(number, number, exact=True)

    raw = str(value).strip()
    if raw.casefold() in MISSING_TOKENS:
        return None

    normalized = (
        raw.replace(",", "")
        .replace("，", "")
        .replace("％", "%")
        .replace(" ", "")
    )

    pct_range = re.fullmatch(
        rf"({NUMBER})%{RANGE_SEPARATORS}({NUMBER})%", normalized
    )
    if pct_range:
        return _result(float(pct_range.group(1)) / 100, float(pct_range.group(2)) / 100)

    numeric_range = re.fullmatch(
        rf"({NUMBER})({UNIT}){RANGE_SEPARATORS}({NUMBER})({UNIT})", normalized
    )
    if numeric_range:
        low = _number_with_unit(numeric_range.group(1), numeric_range.group(2))
        high = _number_with_unit(numeric_range.group(3), numeric_range.group(4))
        return _result(low, high)

    exact_pct = re.fullmatch(rf"({NUMBER})%", normalized)
    if exact_pct:
        number = float(exact_pct.group(1)) / 100
        return _result(number, number, exact=True)

    exact_number = re.fullmatch(rf"({NUMBER})({UNIT})", normalized)
    if exact_number:
        number = _number_with_unit(exact_number.group(1), exact_number.group(2))
        return _result(number, number, exact=True)

    raise ParseError(f"未知指标格式: {value!r}")


def clean_product_id(value: Any) -> str:
    """返回纯数字商品 ID；无法得到 ID 时抛出 ParseError。"""
    if value is None:
        raise ParseError("商品 ID 为空")
    product_id = re.sub(r"[^0-9]", "", str(value))
    if not product_id:
        raise ParseError(f"商品 ID 无法清洗: {value!r}")
    return product_id


def format_excel_date(value: Any, epoch: datetime) -> str | None:
    """转换已知日期值；范围字符串和月份标签保持原样。"""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        converted = from_excel(value, epoch=epoch)
        if isinstance(converted, datetime):
            return converted.strftime("%Y-%m-%d")
        if isinstance(converted, date):
            return converted.isoformat()
        return str(converted)
    return str(value).strip()


def normalize_header(value: Any) -> str:
    """移除表头中的首尾和连续空白；不擅自改写业务名称。"""
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip())


def read_headers(sheet: Any) -> list[str]:
    headers = [normalize_header(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if not any(headers):
        raise ValueError(f"Sheet {sheet.title!r} 第一行没有表头")
    nonempty = [header for header in headers if header]
    duplicates = sorted(name for name, count in Counter(nonempty).items() if count > 1)
    if duplicates:
        raise ValueError(f"Sheet {sheet.title!r} 存在重复表头: {duplicates}")
    return headers


def iter_records(sheet: Any, headers: list[str]) -> Iterable[tuple[int, dict[str, Any]]]:
    """按表头生成记录；完全空白行跳过。"""
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        record = {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
            if header
        }
        yield row_number, record


def _sample_values(sheet: Any, limit_per_column: int = 5) -> dict[str, list[str]]:
    headers = read_headers(sheet)
    samples: dict[str, list[str]] = {header: [] for header in headers if header}
    for _, record in iter_records(sheet, headers):
        for header, value in record.items():
            if value is None or len(samples[header]) >= limit_per_column:
                continue
            rendered = str(value)
            if rendered not in samples[header]:
                samples[header].append(rendered)
    return samples


def audit_workbook(path: str | Path) -> dict[str, Any]:
    """只读审计工作簿结构，并定位缺少缓存结果的公式单元格。"""
    path = Path(path)
    values_wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    formulas_wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        for name in values_wb.sheetnames:
            values_ws = values_wb[name]
            formulas_ws = formulas_wb[name]
            headers = read_headers(values_ws)
            missing_formula_cache: list[str] = []
            for formula_row, value_row in zip(formulas_ws.iter_rows(), values_ws.iter_rows()):
                for formula_cell, value_cell in zip(formula_row, value_row):
                    if (
                        isinstance(formula_cell.value, str)
                        and formula_cell.value.startswith("=")
                        and value_cell.value is None
                    ):
                        missing_formula_cache.append(formula_cell.coordinate)
            sheets.append(
                {
                    "name": name,
                    "rows": values_ws.max_row,
                    "columns": values_ws.max_column,
                    "headers": headers,
                    "samples": _sample_values(values_ws),
                    "formula_cells_without_cached_value": missing_formula_cache[:100],
                    "formula_cache_issue_count": len(missing_formula_cache),
                }
            )
        return {
            "file": path.name,
            "epoch": values_wb.epoch.isoformat(),
            "sheet_count": len(sheets),
            "sheets": sheets,
        }
    finally:
        values_wb.close()
        formulas_wb.close()


def self_test() -> None:
    cases = [
        ("1000 ~ 2500", {"low": 1000.0, "mid": 1750.0, "high": 2500.0}),
        ("1万 ～ 2.5万", {"low": 10000.0, "mid": 17500.0, "high": 25000.0}),
        ("0% 至 1%", {"low": 0.0, "mid": 0.005, "high": 0.01}),
        ("12,500", {"low": 12500.0, "mid": 12500.0, "high": 12500.0, "exact": True}),
        (0, {"low": 0.0, "mid": 0.0, "high": 0.0, "exact": True}),
        ("-", None),
    ]
    for source, expected in cases:
        actual = parse_range(source)
        if actual != expected:
            raise AssertionError(f"parse_range({source!r}) = {actual!r}, expected {expected!r}")
    try:
        parse_range("100~200人")
    except ParseError:
        pass
    else:
        raise AssertionError("未知格式未抛出 ParseError")
    assert clean_product_id("'656821090235") == "656821090235"
    print("self-test passed")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel", nargs="?", help="待审计的 Excel 文件")
    parser.add_argument("--audit-json", help="审计结果 JSON 输出路径")
    parser.add_argument("--self-test", action="store_true", help="运行内置单元测试")
    args = parser.parse_args()

    if args.self_test:
        self_test()
        return
    if not args.excel:
        parser.error("请提供 Excel 文件，或使用 --self-test")

    result = audit_workbook(args.excel)
    rendered = json.dumps(result, ensure_ascii=False, indent=2)
    if args.audit_json:
        Path(args.audit_json).write_text(rendered + "\n", encoding="utf-8")
        print(f"审计结果已保存: {args.audit_json}")
    else:
        print(rendered)


if __name__ == "__main__":
    main()
